"""
Excel Services for Prime Potions ERP
Handles Excel import/export with EXACT Prime Potions template matching
"""
import io
import uuid
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
import hashlib
import json
import re

def generate_id():
    return str(uuid.uuid4())

def get_timestamp():
    return datetime.now(timezone.utc).isoformat()

# =============================================================================
# PRIME POTIONS EXACT TEMPLATE DEFINITIONS
# =============================================================================

# RAW MATERIALS - Sheet: "RAW-MASTER INV"
# Headers with EXACT spelling/spacing/newlines
RAW_MATERIAL_HEADERS = [
    "ITEM CODE",
    "INTERNAL LOT #\n(Item code-YYMMDD)",
    "Ingredient Name",
    "SUPPLIER LOT #",
    "Tracking key",
    "Opening stock",
    "Inventory on hand",
    "EXPIRY / RETEST Date",
    "VENDOR / MANUFACTURER",
    "INCI NAME",
    "Primary Inv Zone",
    "2ND\nInv Zone",
    "CoA\n(Yes/No)",
    "Container Type",
    "Column2",
    "UoM",
    "Notes",
    "Minimum stock",
    "Stock status"
]

RAW_MATERIAL_FIELD_MAP = {
    "ITEM CODE": "sku",
    "INTERNAL LOT #\n(Item code-YYMMDD)": "internal_lot",
    "Ingredient Name": "name",
    "SUPPLIER LOT #": "supplier_lot",
    "Tracking key": "tracking_key",
    "Opening stock": "opening_stock",
    "Inventory on hand": "on_hand",
    "EXPIRY / RETEST Date": "expiry_date",
    "VENDOR / MANUFACTURER": "manufacturer",
    "INCI NAME": "inci_name",
    "Primary Inv Zone": "location",
    "2ND\nInv Zone": "secondary_location",
    "CoA\n(Yes/No)": "coa_status",
    "Container Type": "container_type",
    "Column2": "custom_column2",
    "UoM": "unit_of_measure",
    "Notes": "notes",
    "Minimum stock": "min_stock_level",
    "Stock status": "stock_status"
}

# PACKAGING - Sheet: "Master inventory-Packaging"
PACKAGING_HEADERS = [
    "Item Name",
    "sub category",
    "category",
    "Client",
    "Supplier",
    "Size or Specs",
    "UOM",
    "Opening  Stock",
    "On Hand",
    "Active",
    "Storage location",
    "Minimum Stock",
    "Stock Status"
]

PACKAGING_FIELD_MAP = {
    "Item Name": "name",
    "sub category": "sub_category",
    "category": "category",
    "Client": "client",
    "Supplier": "supplier",
    "Size or Specs": "size_specs",
    "UOM": "unit_of_measure",
    "Opening  Stock": "opening_stock",
    "On Hand": "on_hand",
    "Active": "is_active",
    "Storage location": "location",
    "Minimum Stock": "min_stock_level",
    "Stock Status": "stock_status"
}

# BATCHING - Sheet: "Batching Sheet" - Header row 4
# Columns A-N in EXACT order including blanks
BATCHING_HEADERS = [
    "Ingredient Formula",      # A
    "Inv Loc.",                 # B
    "Qty Required",             # C
    "Add Order",                # D
    "Added",                    # E
    "Kg Sum",                   # F
    "Process Notes",            # G
    "Batch Notes",              # H
    "",                         # I - BLANK
    "Qty on Hand (kg)",         # J
    "ENTER % QTY HERE",         # K
    "",                         # L - BLANK
    "",                         # M - BLANK
]

BATCHING_FIELD_MAP = {
    "Ingredient Formula": "ingredient_name",
    "Inv Loc.": "location",
    "Qty Required": "qty_required",
    "Add Order": "add_order",
    "Added": "actual_qty",
    "Kg Sum": "running_total",
    "Process Notes": "process_notes",
    "Batch Notes": "batch_notes",
    "Qty on Hand (kg)": "qty_on_hand",
    "ENTER % QTY HERE": "percent_qty"
}


class PrimePotionsExcelService:
    """
    Excel service that generates/parses files matching Prime Potions' exact templates
    """
    
    # Styles
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    HEADER_FILL = PatternFill(start_color="0F5132", end_color="0F5132", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def generate_raw_materials_excel(items: List[Dict], inventory_data: List[Dict]) -> bytes:
        """
        Generate Raw Materials Excel with exact Prime Potions headers
        Sheet name: RAW-MASTER INV
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "RAW-MASTER INV"
        
        # Write headers
        for col_idx, header in enumerate(RAW_MATERIAL_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = PrimePotionsExcelService.HEADER_FONT
            cell.fill = PrimePotionsExcelService.HEADER_FILL
            cell.border = PrimePotionsExcelService.THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        # Create inventory lookup by item_id
        inv_by_item = {}
        for inv in inventory_data:
            item_id = inv.get("item_id")
            if item_id not in inv_by_item:
                inv_by_item[item_id] = []
            inv_by_item[item_id].append(inv)
        
        # Write data rows
        row_idx = 2
        for item in items:
            item_id = item.get("id")
            item_inv = inv_by_item.get(item_id, [{}])
            
            # If multiple lots, create row per lot; otherwise single row
            if item_inv:
                for lot_data in item_inv:
                    ws.cell(row=row_idx, column=1, value=item.get("sku", ""))
                    ws.cell(row=row_idx, column=2, value=lot_data.get("lot_number", ""))
                    ws.cell(row=row_idx, column=3, value=item.get("name", ""))
                    ws.cell(row=row_idx, column=4, value=item.get("supplier_lot", ""))
                    ws.cell(row=row_idx, column=5, value=item.get("tracking_key", ""))
                    ws.cell(row=row_idx, column=6, value=item.get("opening_stock", 0))
                    ws.cell(row=row_idx, column=7, value=lot_data.get("quantity_on_hand", 0))
                    ws.cell(row=row_idx, column=8, value=item.get("expiry_date", ""))
                    ws.cell(row=row_idx, column=9, value=item.get("manufacturer", ""))
                    ws.cell(row=row_idx, column=10, value=item.get("inci_name", ""))
                    ws.cell(row=row_idx, column=11, value=item.get("location", ""))
                    ws.cell(row=row_idx, column=12, value=item.get("secondary_location", ""))
                    ws.cell(row=row_idx, column=13, value=item.get("coa_status", ""))
                    ws.cell(row=row_idx, column=14, value=item.get("container_type", ""))
                    ws.cell(row=row_idx, column=15, value=item.get("custom_column2", ""))
                    ws.cell(row=row_idx, column=16, value=item.get("unit_of_measure", "KG"))
                    ws.cell(row=row_idx, column=17, value=item.get("notes", ""))
                    ws.cell(row=row_idx, column=18, value=item.get("min_stock_level", 0))
                    
                    # Stock status calculation
                    on_hand = lot_data.get("quantity_on_hand", 0) or 0
                    min_stock = item.get("min_stock_level", 0) or 0
                    if on_hand <= 0:
                        status = "OUT OF STOCK"
                    elif min_stock > 0 and on_hand < min_stock:
                        status = "LOW STOCK"
                    else:
                        status = "IN STOCK"
                    ws.cell(row=row_idx, column=19, value=status)
                    row_idx += 1
            else:
                # No inventory data, still show item
                ws.cell(row=row_idx, column=1, value=item.get("sku", ""))
                ws.cell(row=row_idx, column=3, value=item.get("name", ""))
                ws.cell(row=row_idx, column=9, value=item.get("manufacturer", ""))
                ws.cell(row=row_idx, column=10, value=item.get("inci_name", ""))
                ws.cell(row=row_idx, column=11, value=item.get("location", ""))
                ws.cell(row=row_idx, column=16, value=item.get("unit_of_measure", "KG"))
                ws.cell(row=row_idx, column=18, value=item.get("min_stock_level", 0))
                ws.cell(row=row_idx, column=19, value="NO INVENTORY")
                row_idx += 1
        
        # Auto-width columns
        for col_idx in range(1, len(RAW_MATERIAL_HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def generate_packaging_excel(items: List[Dict], inventory_data: List[Dict]) -> bytes:
        """
        Generate Packaging Excel with exact Prime Potions headers
        Sheet name: Master inventory-Packaging
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Master inventory-Packaging"
        
        # Write headers
        for col_idx, header in enumerate(PACKAGING_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = PrimePotionsExcelService.HEADER_FONT
            cell.fill = PrimePotionsExcelService.HEADER_FILL
            cell.border = PrimePotionsExcelService.THIN_BORDER
        
        # Create inventory lookup
        inv_by_item = {}
        for inv in inventory_data:
            item_id = inv.get("item_id")
            if item_id not in inv_by_item:
                inv_by_item[item_id] = {"total": 0}
            inv_by_item[item_id]["total"] += inv.get("quantity_on_hand", 0)
        
        # Write data rows
        row_idx = 2
        for item in items:
            item_id = item.get("id")
            inv_total = inv_by_item.get(item_id, {}).get("total", 0)
            
            ws.cell(row=row_idx, column=1, value=item.get("name", ""))
            ws.cell(row=row_idx, column=2, value=item.get("sub_category", ""))
            ws.cell(row=row_idx, column=3, value=item.get("category", ""))
            ws.cell(row=row_idx, column=4, value=item.get("client", ""))
            ws.cell(row=row_idx, column=5, value=item.get("supplier", ""))
            ws.cell(row=row_idx, column=6, value=item.get("size_specs", ""))
            ws.cell(row=row_idx, column=7, value=item.get("unit_of_measure", "EA"))
            ws.cell(row=row_idx, column=8, value=item.get("opening_stock", 0))
            ws.cell(row=row_idx, column=9, value=inv_total)
            ws.cell(row=row_idx, column=10, value="Yes" if item.get("is_active", True) else "No")
            ws.cell(row=row_idx, column=11, value=item.get("location", ""))
            ws.cell(row=row_idx, column=12, value=item.get("min_stock_level", 0))
            
            # Stock status
            min_stock = item.get("min_stock_level", 0) or 0
            if inv_total <= 0:
                status = "OUT OF STOCK"
            elif min_stock > 0 and inv_total < min_stock:
                status = "LOW STOCK"
            else:
                status = "IN STOCK"
            ws.cell(row=row_idx, column=13, value=status)
            row_idx += 1
        
        # Auto-width
        for col_idx in range(1, len(PACKAGING_HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def generate_batching_template(
        batch_info: Dict[str, Any],
        formula_lines: List[Dict[str, Any]],
        inventory_lookup: List[Dict[str, Any]]
    ) -> bytes:
        """
        Generate Batching Excel with exact Prime Potions format
        - Sheet: "Batching Sheet" with header row 4
        - Sheet: "Do not change - Import range fr" for VLOOKUP support
        """
        wb = Workbook()
        
        # ============ BATCHING SHEET ============
        ws_batch = wb.active
        ws_batch.title = "Batching Sheet"
        
        # Title rows (1-3)
        ws_batch.cell(row=1, column=1, value=f"BATCH: {batch_info.get('batch_code', '')}")
        ws_batch.cell(row=1, column=1).font = Font(bold=True, size=14)
        product_line = f"Product: {batch_info.get('product_name', batch_info.get('formula_name', ''))}"
        if batch_info.get('formula_revision'):
            product_line += f" ({batch_info['formula_revision']})"
        ws_batch.cell(row=2, column=1, value=product_line)
        ws_batch.cell(row=2, column=3, value=f"Target Size: {batch_info.get('planned_qty', '')} {batch_info.get('batch_unit', 'KG')}")
        ws_batch.cell(row=3, column=1, value=f"Date: {batch_info.get('batch_date', datetime.now().strftime('%Y-%m-%d'))}")
        note_cell = ws_batch.cell(
            row=3, column=4,
            value="To add an ingredient not listed below: insert a new row directly above FINISH WT (don't leave a blank row) with the name and the Added quantity."
        )
        note_cell.font = Font(italic=True, size=9, color="777777")

        # Header row 4 - EXACT columns A-N
        for col_idx, header in enumerate(BATCHING_HEADERS, 1):
            cell = ws_batch.cell(row=4, column=col_idx, value=header)
            cell.font = PrimePotionsExcelService.HEADER_FONT
            cell.fill = PrimePotionsExcelService.HEADER_FILL
            cell.border = PrimePotionsExcelService.THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        
        # Build inventory lookup dict
        inv_lookup = {}
        for inv in inventory_lookup:
            name = inv.get("name", "").strip()
            if name:
                if name not in inv_lookup:
                    inv_lookup[name] = {"location": inv.get("location", ""), "qty": 0}
                inv_lookup[name]["qty"] += inv.get("quantity_on_hand", 0)
        
        # Write formula lines starting row 5
        row_idx = 5
        running_total = 0
        for line in formula_lines:
            ingredient_name = line.get("ingredient_display_name", line.get("name", ""))
            qty_required = line.get("default_qty_required", line.get("qty_required", 0)) or 0
            
            # Get on-hand from lookup
            inv_data = inv_lookup.get(ingredient_name, {"location": "", "qty": 0})
            
            running_total += qty_required
            
            ws_batch.cell(row=row_idx, column=1, value=ingredient_name)  # A: Ingredient Formula
            ws_batch.cell(row=row_idx, column=2, value=inv_data["location"])  # B: Inv Loc.
            ws_batch.cell(row=row_idx, column=3, value=qty_required)  # C: Qty Required
            ws_batch.cell(row=row_idx, column=4, value=line.get("add_order", ""))  # D: Add Order
            ws_batch.cell(row=row_idx, column=5, value="")  # E: Added (user fills)
            ws_batch.cell(row=row_idx, column=6, value=running_total)  # F: Kg Sum
            ws_batch.cell(row=row_idx, column=7, value=line.get("process_notes", ""))  # G: Process Notes
            ws_batch.cell(row=row_idx, column=8, value=line.get("batch_notes", ""))  # H: Batch Notes
            ws_batch.cell(row=row_idx, column=9, value="")  # I: BLANK
            ws_batch.cell(row=row_idx, column=10, value=inv_data["qty"])  # J: Qty on Hand (kg)
            ws_batch.cell(row=row_idx, column=11, value=line.get("percent", ""))  # K: % QTY
            # L, M, N are blank or info columns
            
            row_idx += 1
        
        # Add FINISH WT row
        ws_batch.cell(row=row_idx, column=1, value="FINISH WT")
        ws_batch.cell(row=row_idx, column=1).font = Font(bold=True)
        ws_batch.cell(row=row_idx, column=5, value="")  # User enters actual finish weight
        
        # Column widths
        col_widths = [25, 12, 12, 10, 10, 10, 20, 20, 5, 15, 15, 5, 5, 40]
        for col_idx, width in enumerate(col_widths, 1):
            ws_batch.column_dimensions[get_column_letter(col_idx)].width = width
        
        # ============ HELPER SHEET FOR VLOOKUP ============
        ws_lookup = wb.create_sheet("Do not change - Import range fr")
        
        # Headers for lookup: A=Ingredient, B=SKU, C=Location, D=UOM, E=Qty on Hand
        lookup_headers = ["Ingredient Formula", "SKU", "Inv Loc", "UOM", "Qty on Hand (kg)"]
        for col_idx, header in enumerate(lookup_headers, 1):
            ws_lookup.cell(row=1, column=col_idx, value=header)
            ws_lookup.cell(row=1, column=col_idx).font = Font(bold=True)
        
        # Populate lookup data
        row_idx = 2
        for inv in inventory_lookup:
            ws_lookup.cell(row=row_idx, column=1, value=inv.get("name", ""))
            ws_lookup.cell(row=row_idx, column=2, value=inv.get("sku", ""))
            ws_lookup.cell(row=row_idx, column=3, value=inv.get("location", ""))
            ws_lookup.cell(row=row_idx, column=4, value=inv.get("unit_of_measure", "KG"))
            ws_lookup.cell(row=row_idx, column=5, value=inv.get("quantity_on_hand", 0))
            row_idx += 1
        
        # Hide the lookup sheet (optional - keeps it accessible but not prominent)
        ws_lookup.sheet_state = 'hidden'
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def parse_batching_upload(file_content: bytes) -> Dict[str, Any]:
        """
        Parse uploaded batching sheet (Prime Potions format)
        Reads from row 5 until blank Ingredient Formula
        """
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        
        result = {
            "batch_info": {},
            "ingredients": [],
            "finish_weight": None,
            "warnings": [],
            "errors": []
        }
        
        # Find Batching Sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "batch" in name.lower():
                sheet_name = name
                break
        
        if not sheet_name:
            result["errors"].append("Could not find batching sheet")
            return result
        
        ws = wb[sheet_name]
        
        # Parse title info from rows 1-3
        batch_code_cell = ws.cell(row=1, column=1).value or ""
        if "BATCH:" in str(batch_code_cell):
            result["batch_info"]["batch_code"] = batch_code_cell.replace("BATCH:", "").strip()
        
        product_cell = ws.cell(row=2, column=1).value or ""
        if "Product:" in str(product_cell):
            result["batch_info"]["product_name"] = product_cell.replace("Product:", "").strip()
        
        # Parse ingredients starting from row 5. Tolerate a stray blank row (e.g. left over
        # from inserting a new ingredient row) instead of silently stopping there and
        # missing everything below it, including FINISH WT.
        row_idx = 5
        consecutive_blanks = 0
        while True:
            ingredient = ws.cell(row=row_idx, column=1).value
            if not ingredient or str(ingredient).strip() == "":
                consecutive_blanks += 1
                if consecutive_blanks >= 3:
                    break
                row_idx += 1
                continue
            consecutive_blanks = 0

            ingredient_str = str(ingredient).strip()
            
            # Check for FINISH WT row
            if ingredient_str.upper() == "FINISH WT":
                result["finish_weight"] = ws.cell(row=row_idx, column=5).value
                row_idx += 1
                continue
            
            # Parse ingredient row
            ing_data = {
                "ingredient_name": ingredient_str,
                "location": ws.cell(row=row_idx, column=2).value or "",
                "qty_required": ws.cell(row=row_idx, column=3).value or 0,
                "add_order": ws.cell(row=row_idx, column=4).value or "",
                "actual_qty": ws.cell(row=row_idx, column=5).value,  # Added column
                "kg_sum": ws.cell(row=row_idx, column=6).value or 0,
                "process_notes": ws.cell(row=row_idx, column=7).value or "",
                "batch_notes": ws.cell(row=row_idx, column=8).value or "",
                "qty_on_hand": ws.cell(row=row_idx, column=10).value or 0,
                "percent_qty": ws.cell(row=row_idx, column=11).value or ""
            }
            
            # Convert qty values to float
            try:
                ing_data["qty_required"] = float(ing_data["qty_required"]) if ing_data["qty_required"] else 0
            except:
                ing_data["qty_required"] = 0
            
            try:
                if ing_data["actual_qty"] is not None and ing_data["actual_qty"] != "":
                    ing_data["actual_qty"] = float(ing_data["actual_qty"])
                else:
                    ing_data["actual_qty"] = None
            except:
                ing_data["actual_qty"] = None
            
            result["ingredients"].append(ing_data)
            row_idx += 1
            
            if row_idx > 500:  # Safety limit
                break
        
        # Convert finish weight
        if result["finish_weight"]:
            try:
                result["finish_weight"] = float(result["finish_weight"])
            except:
                result["warnings"].append(f"Could not parse finish weight: {result['finish_weight']}")
                result["finish_weight"] = None
        
        return result
    
    @staticmethod
    def parse_raw_materials_import(file_content: bytes) -> Dict[str, Any]:
        """Parse Raw Materials Excel import"""
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        
        result = {
            "items": [],
            "errors": [],
            "warnings": []
        }
        
        # Find the right sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "raw" in name.lower() and "master" in name.lower():
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
        
        ws = wb[sheet_name]
        
        # Get headers from row 1
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val) if val else f"col_{col}")
        
        # Parse data rows
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_idx, header in enumerate(headers, 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    has_data = True
                
                # Map header to field
                field = RAW_MATERIAL_FIELD_MAP.get(header)
                if field:
                    row_data[field] = val
                else:
                    # Store as custom field
                    row_data[f"custom_{header}"] = val
            
            if has_data and (row_data.get("sku") or row_data.get("name")):
                result["items"].append(row_data)
        
        return result
    
    @staticmethod
    def parse_packaging_import(file_content: bytes) -> Dict[str, Any]:
        """Parse Packaging Excel import"""
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        
        result = {
            "items": [],
            "errors": [],
            "warnings": []
        }
        
        # Find the right sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "packaging" in name.lower() or "master" in name.lower():
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
        
        ws = wb[sheet_name]
        
        # Get headers
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val) if val else f"col_{col}")
        
        # Parse data rows
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_idx, header in enumerate(headers, 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    has_data = True
                
                field = PACKAGING_FIELD_MAP.get(header)
                if field:
                    row_data[field] = val
                else:
                    row_data[f"custom_{header}"] = val
            
            if has_data and row_data.get("name"):
                result["items"].append(row_data)
        
        return result


class ExcelTemplateConfig:
    """
    Manages configurable Excel template mappings
    Allows Admin to adjust column parsing without code changes
    """
    
    DEFAULT_BATCHING_CONFIG = {
        "template_key": "batching_prime_potions_v1",
        "sheet_name": "Batching Sheet",
        "header_row": 4,
        "columns": [
            {"key": "ingredient_name", "header": "Ingredient Formula", "col": "A", "required": True, "enabled": True},
            {"key": "location", "header": "Inv Loc.", "col": "B", "required": False, "enabled": True},
            {"key": "qty_required", "header": "Qty Required", "col": "C", "required": True, "enabled": True},
            {"key": "add_order", "header": "Add Order", "col": "D", "required": False, "enabled": True},
            {"key": "actual_qty", "header": "Added", "col": "E", "required": False, "enabled": True},
            {"key": "kg_sum", "header": "Kg Sum", "col": "F", "required": False, "enabled": True},
            {"key": "process_notes", "header": "Process Notes", "col": "G", "required": False, "enabled": True},
            {"key": "batch_notes", "header": "Batch Notes", "col": "H", "required": False, "enabled": True},
            {"key": "blank_1", "header": "", "col": "I", "required": False, "enabled": False},
            {"key": "qty_on_hand", "header": "Qty on Hand (kg)", "col": "J", "required": False, "enabled": True},
            {"key": "percent_qty", "header": "ENTER % QTY HERE", "col": "K", "required": False, "enabled": True},
            {"key": "blank_2", "header": "", "col": "L", "required": False, "enabled": False},
            {"key": "blank_3", "header": "", "col": "M", "required": False, "enabled": False},
        ]
    }
    
    @staticmethod
    def get_config_export_excel() -> bytes:
        """Export template config as Excel for Admin editing"""
        wb = Workbook()
        ws = wb.active
        ws.title = "BatchingColumns"
        
        headers = ["col_letter", "header_text", "internal_key", "enabled", "required", "notes"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
        
        for row_idx, col_cfg in enumerate(ExcelTemplateConfig.DEFAULT_BATCHING_CONFIG["columns"], 2):
            ws.cell(row=row_idx, column=1, value=col_cfg["col"])
            ws.cell(row=row_idx, column=2, value=col_cfg["header"])
            ws.cell(row=row_idx, column=3, value=col_cfg["key"])
            ws.cell(row=row_idx, column=4, value="Yes" if col_cfg["enabled"] else "No")
            ws.cell(row=row_idx, column=5, value="Yes" if col_cfg["required"] else "No")
            ws.cell(row=row_idx, column=6, value="")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()



# Standard field mappings for generic Excel files
DEFAULT_RAW_MATERIAL_MAPPINGS = {
    "item_code": ["ITEM CODE", "Item Code", "SKU", "Code", "ItemCode"],
    "name": ["Ingredient Name", "Name", "INCI NAME", "Material Name", "Item Name"],
    "supplier": ["SUPPLIER", "Supplier", "Vendor"],
    "lot_number": ["LOT #", "Lot Number", "LOT", "Lot #", "Lot"],
    "expiry_date": ["EXPIRY / RETEST Date", "Expiry Date", "Expiry", "Retest Date"],
    "manufacturer": ["VENDOR / MANUFACTURER", "Manufacturer", "Vendor / Manufacturer"],
    "inci_name": ["INCI NAME", "INCI", "Scientific Name"],
    "location": ["Primary Inv Zone", "Storage location", "Location", "Zone"],
    "quantity_on_hand": ["Inventory on Hand", "On Hand", "QTY", "Quantity", "Stock"],
    "container_type": ["Container Type", "Container", "Packaging"],
    "uom": ["UoM", "UOM", "Unit", "Unit of Measure"],
    "notes": ["Notes", "Comments", "Remarks"],
    "minimum_stock": ["Safety Stock", "Minimum Stock", "Min Stock", "Reorder Point"],
    "category": ["Class", "Category", "sub category", "Type"],
    "cost_per_unit": ["Cost/unit", "Cost", "Unit Cost", "Price"]
}

DEFAULT_PACKAGING_MAPPINGS = {
    "item_code": ["Item Code", "SKU", "Code"],
    "name": ["Item Name", "Name", "Description"],
    "category": ["category", "Category", "Type"],
    "sub_category": ["sub category", "Sub Category", "Subcategory"],
    "client": ["Client", "Customer"],
    "supplier": ["Supplier", "Vendor"],
    "size_specs": ["Size or Specs", "Size", "Specs", "Specifications"],
    "uom": ["UOM", "UoM", "Unit"],
    "quantity_on_hand": ["On Hand", "Stock", "Quantity"],
    "location": ["Storage location", "Location"],
    "minimum_stock": ["Minimum Stock", "Min Stock"],
    "stock_status": ["Stock Status", "Status"]
}


def fuzzy_match_column(column_name: str, mappings: Dict[str, List[str]]) -> Optional[str]:
    """Find the best matching standard field for a column name"""
    col_lower = column_name.lower().strip()
    
    for standard_field, aliases in mappings.items():
        for alias in aliases:
            if alias.lower() == col_lower:
                return standard_field
    
    # Partial match
    for standard_field, aliases in mappings.items():
        for alias in aliases:
            if alias.lower() in col_lower or col_lower in alias.lower():
                return standard_field
    
    return None


class ExcelService:
    """Generic Excel service for analyzing and parsing workbooks"""
    
    @staticmethod
    def analyze_workbook(content: bytes) -> Dict[str, Any]:
        """Analyze an Excel workbook and return its structure"""
        wb = load_workbook(io.BytesIO(content), data_only=True)
        
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Get headers from first row
            headers = []
            for col in range(1, min(ws.max_column + 1, 50)):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers.append(str(val))
            
            # Count data rows
            row_count = 0
            for row in range(2, min(ws.max_row + 1, 1000)):
                if ws.cell(row=row, column=1).value:
                    row_count += 1
            
            sheets.append({
                "name": sheet_name,
                "headers": headers,
                "row_count": row_count
            })
        
        return {"sheets": sheets}
    
    @staticmethod
    def suggest_mappings(headers: List[str], mapping_type: str = "raw_material") -> Dict[str, str]:
        """Suggest field mappings for column headers"""
        mappings_dict = DEFAULT_RAW_MATERIAL_MAPPINGS if mapping_type == "raw_material" else DEFAULT_PACKAGING_MAPPINGS
        
        suggestions = {}
        for header in headers:
            matched = fuzzy_match_column(header, mappings_dict)
            if matched:
                suggestions[header] = matched
        
        return suggestions
    
    @staticmethod
    def parse_excel_to_records(content: bytes, sheet_name: str, field_mappings: Dict[str, str]) -> List[Dict[str, Any]]:
        """Parse Excel sheet to list of records using field mappings"""
        wb = load_workbook(io.BytesIO(content), data_only=True)
        
        if sheet_name not in wb.sheetnames:
            return []
        
        ws = wb[sheet_name]
        
        # Get headers
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val) if val else f"col_{col}")
        
        records = []
        for row_idx in range(2, ws.max_row + 1):
            record = {}
            has_data = False
            
            for col_idx, header in enumerate(headers, 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    has_data = True
                
                # Map to standard field if mapping exists
                field_name = field_mappings.get(header, header)
                record[field_name] = val
            
            if has_data:
                records.append(record)
        
        return records
    
    @staticmethod
    def generate_master_data_template(template_type: str) -> bytes:
        """Generate a template Excel file for data import"""
        wb = Workbook()
        ws = wb.active
        
        if template_type == "raw_materials":
            ws.title = "Raw Materials"
            headers = ["SKU", "Name", "Category", "UOM", "Manufacturer", "INCI Name", "Location", "Min Stock", "Notes"]
        elif template_type == "packaging":
            ws.title = "Packaging"
            headers = ["SKU", "Name", "Category", "Sub Category", "Supplier", "Size/Specs", "UOM", "Location", "Min Stock"]
        else:
            ws.title = "Inventory Receipt"
            headers = ["SKU", "Name", "Lot Number", "Quantity", "UOM", "Location", "Expiry Date", "Notes"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class ImportPreviewService:
    """Service for generating import previews"""
    
    @staticmethod
    async def generate_preview(records: List[Dict], existing_items: Dict, key_field: str) -> Dict[str, Any]:
        """Generate a preview of what changes would be made by an import"""
        preview = {
            "to_create": [],
            "to_update": [],
            "unchanged": [],
            "errors": [],
            "total_records": len(records)
        }
        
        for record in records:
            key = record.get(key_field, record.get("sku", ""))
            if not key:
                preview["errors"].append({"record": record, "error": f"Missing {key_field}"})
                continue
            
            existing = existing_items.get(key)
            
            if existing:
                # Check if anything changed
                changed_fields = []
                for field, value in record.items():
                    if field.startswith("_"):
                        continue
                    existing_value = existing.get(field)
                    if value != existing_value and value is not None:
                        changed_fields.append({
                            "field": field,
                            "old": existing_value,
                            "new": value
                        })
                
                if changed_fields:
                    preview["to_update"].append({
                        "key": key,
                        "changes": changed_fields
                    })
                else:
                    preview["unchanged"].append(key)
            else:
                preview["to_create"].append({
                    "key": key,
                    "record": {k: v for k, v in record.items() if not k.startswith("_")}
                })
        
        return preview
